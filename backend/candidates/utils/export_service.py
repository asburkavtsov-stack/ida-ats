import io
import os
import urllib.request
import zipfile
import traceback
import logging
from datetime import datetime
from typing import Dict, List, Any, Optional

from django.http import HttpResponse
from django.utils import timezone

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib.fonts import addMapping
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

logger = logging.getLogger(__name__)

# ─────────────────────────────────────────────────────────────
# ШРИФТИ — завантажуємо DejaVu якщо немає локально
# ─────────────────────────────────────────────────────────────

def _ensure_cyrillic_font():
    """Переконується, що шрифт з кирилицею доступний для ReportLab"""
    if not REPORTLAB_AVAILABLE:
        return 'Helvetica'

    # Перевіряємо чи вже зареєстровано
    try:
        pdfmetrics.getFont('DejaVuSans')
        return 'DejaVuSans'
    except:
        pass

    # Спробуємо знайти локально
    local_paths = [
        '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
        '/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf',
        '/usr/share/fonts/truetype/liberation2/LiberationSans-Regular.ttf',
        '/usr/share/fonts/truetype/freefont/FreeSans.ttf',
        '/usr/share/fonts/truetype/ubuntu/Ubuntu-R.ttf',
        'C:/Windows/Fonts/arial.ttf',
        '/Library/Fonts/Arial.ttf',
    ]

    for path in local_paths:
        if os.path.exists(path):
            try:
                pdfmetrics.registerFont(TTFont('DejaVuSans', path))
                addMapping('DejaVuSans', 0, 0, 'DejaVuSans')
                logger.info(f"Font loaded from {path}")
                return 'DejaVuSans'
            except Exception as e:
                logger.warning(f"Failed to load font from {path}: {e}")
                continue

    # Якщо не знайшли локально — завантажуємо DejaVu
    font_dir = os.path.join(os.path.dirname(__file__), 'fonts')
    os.makedirs(font_dir, exist_ok=True)
    font_path = os.path.join(font_dir, 'DejaVuSans.ttf')

    if not os.path.exists(font_path):
        try:
            # Завантажуємо DejaVu Sans з офіційного джерела
            url = "https://github.com/dejavu-fonts/dejavu-fonts/raw/master/ttf/DejaVuSans.ttf"
            urllib.request.urlretrieve(url, font_path)
            logger.info("Downloaded DejaVuSans.ttf")
        except Exception as e:
            logger.error(f"Failed to download font: {e}")
            # Спробуємо завантажити через альтернативний URL
            try:
                url = "https://raw.githubusercontent.com/dejavu-fonts/dejavu-fonts/master/ttf/DejaVuSans.ttf"
                urllib.request.urlretrieve(url, font_path)
                logger.info("Downloaded DejaVuSans.ttf (alt)")
            except Exception as e2:
                logger.error(f"Failed alt download: {e2}")
                return 'Helvetica'

    try:
        pdfmetrics.registerFont(TTFont('DejaVuSans', font_path))
        addMapping('DejaVuSans', 0, 0, 'DejaVuSans')
        return 'DejaVuSans'
    except Exception as e:
        logger.error(f"Failed to register downloaded font: {e}")
        return 'Helvetica'


# ─────────────────────────────────────────────────────────────
# КОЛЬОРОВА СХЕМА
# ─────────────────────────────────────────────────────────────

COLORS = {
    'primary': colors.HexColor('#1a1a2e'),
    'secondary': colors.HexColor('#16213e'),
    'accent': colors.HexColor('#0f3460'),
    'success': colors.HexColor('#16a34a'),
    'warning': colors.HexColor('#eab308'),
    'danger': colors.HexColor('#dc2626'),
    'text': colors.HexColor('#1f2937'),
    'muted': colors.HexColor('#6b7280'),
    'bg': colors.HexColor('#f8fafc'),
    'border': colors.HexColor('#e2e8f0'),
    'white': colors.HexColor('#ffffff'),
    'header_bg': colors.HexColor('#1e293b'),
}


class ExportService:
    """Сервіс для експорту даних у Excel та PDF"""

    @classmethod
    def _get_styles(cls):
        """Повертає стилі для PDF"""
        font_name = _ensure_cyrillic_font()
        styles = getSampleStyleSheet()

        # Оновлюємо всі стандартні стилі
        for style_name in styles.byName:
            styles[style_name].fontName = font_name

        # Кастомні стилі
        styles.add(ParagraphStyle(
            name='ReportTitle',
            parent=styles['Title'],
            fontName=font_name,
            fontSize=22,
            textColor=COLORS['primary'],
            spaceAfter=6,
            alignment=0,  # Left
        ))

        styles.add(ParagraphStyle(
            name='ReportSubtitle',
            parent=styles['Normal'],
            fontName=font_name,
            fontSize=9,
            textColor=COLORS['muted'],
            spaceAfter=20,
        ))

        styles.add(ParagraphStyle(
            name='SectionHeader',
            parent=styles['Heading2'],
            fontName=font_name,
            fontSize=13,
            textColor=COLORS['primary'],
            spaceBefore=16,
            spaceAfter=10,
            borderPadding=5,
        ))

        styles.add(ParagraphStyle(
            name='TableHeader',
            parent=styles['Normal'],
            fontName=font_name,
            fontSize=8,
            textColor=COLORS['white'],
            alignment=1,  # Center
        ))

        styles.add(ParagraphStyle(
            name='TableCell',
            parent=styles['Normal'],
            fontName=font_name,
            fontSize=8,
            textColor=COLORS['text'],
            leading=11,
        ))

        styles.add(ParagraphStyle(
            name='TableCellRight',
            parent=styles['Normal'],
            fontName=font_name,
            fontSize=8,
            textColor=COLORS['text'],
            alignment=2,  # Right
            leading=11,
        ))

        styles.add(ParagraphStyle(
            name='MetricLabel',
            parent=styles['Normal'],
            fontName=font_name,
            fontSize=8,
            textColor=COLORS['muted'],
            alignment=1,
        ))

        styles.add(ParagraphStyle(
            name='MetricValue',
            parent=styles['Normal'],
            fontName=font_name,
            fontSize=16,
            textColor=COLORS['primary'],
            alignment=1,
        ))

        styles.add(ParagraphStyle(
            name='SmallNote',
            parent=styles['Normal'],
            fontName=font_name,
            fontSize=7,
            textColor=COLORS['muted'],
        ))

        return styles

    # ═══════════════════════════════════════════════════════════
    # TIME-TO-HIRE EXCEL
    # ═══════════════════════════════════════════════════════════

    @classmethod
    def export_time_to_hire_excel(cls, time_data, statistics, filters):
        wb = openpyxl.Workbook()
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1e293b', end_color='1e293b', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        number_alignment = Alignment(horizontal='right', vertical='center')
        center_alignment = Alignment(horizontal='center', vertical='center')

        # Загальна статистика
        ws_summary = wb.active
        ws_summary.title = "Загальна статистика"
        ws_summary.merge_cells('A1:D1')
        ws_summary['A1'] = "Звіт Time-to-Hire"
        ws_summary['A1'].font = Font(name='Arial', size=18, bold=True, color='1e293b')
        ws_summary['A1'].alignment = center_alignment

        ws_summary['A3'] = "Дата генерації:"
        ws_summary['B3'] = timezone.now().strftime('%d.%m.%Y %H:%M:%S')
        ws_summary['A3'].font = Font(bold=True)

        row = 5
        ws_summary[f'A{row}'] = "Застосовані фільтри:"
        ws_summary[f'A{row}'].font = Font(bold=True, size=11, color='1e293b')
        row += 1
        for key, value in filters.items():
            if value:
                ws_summary[f'A{row}'] = f"• {key}:"
                ws_summary[f'B{row}'] = str(value)
                row += 1

        row += 1
        ws_summary[f'A{row}'] = "Основні метрики"
        ws_summary[f'A{row}'].font = Font(bold=True, size=13, color='1e293b')
        row += 1
        for key, value in [("Середній час до офферу (дні)", statistics.get('overall_avg', '—')),
                           ("Медіана (дні)", statistics.get('median', '—')),
                           ("Всього офферів", statistics.get('total_offers', 0))]:
            ws_summary[f'A{row}'] = key
            ws_summary[f'B{row}'] = value
            ws_summary[f'A{row}'].font = Font(bold=True)
            row += 1

        if statistics.get('distribution'):
            row += 1
            ws_summary[f'A{row}'] = "Розподіл по швидкості"
            ws_summary[f'A{row}'].font = Font(bold=True, size=13, color='1e293b')
            row += 1
            for dist in statistics['distribution']:
                ws_summary[f'A{row}'] = dist['range']
                ws_summary[f'B{row}'] = f"{dist['count']} ({dist['percentage']}%)"
                row += 1

        # По вакансіях
        ws_vacancies = wb.create_sheet("По вакансіях")
        vacancy_headers = ["Вакансія", "Середній час (дні)", "Медіана (дні)", "Кількість офферів", "Мін. час", "Макс. час"]
        for col, header in enumerate(vacancy_headers, 1):
            cell = ws_vacancies.cell(row=1, column=col, value=header)
            cell.font = header_font; cell.fill = header_fill; cell.alignment = header_alignment; cell.border = thin_border
        for row_idx, vacancy in enumerate(statistics.get('by_vacancy', []), 2):
            for col_idx, key in enumerate(['vacancy_title', 'avg_days', 'median_days', 'offers_count', 'min_days', 'max_days'], 1):
                cell = ws_vacancies.cell(row=row_idx, column=col_idx, value=vacancy[key])
                cell.border = thin_border
                if col_idx > 1:
                    cell.alignment = number_alignment
        for col in range(1, 7):
            ws_vacancies.column_dimensions[get_column_letter(col)].width = 20

        # По періодах
        if statistics.get('by_period'):
            ws_periods = wb.create_sheet("По періодах")
            for col, header in enumerate(["Період", "Середній час (дні)", "Кількість офферів"], 1):
                cell = ws_periods.cell(row=1, column=col, value=header)
                cell.font = header_font; cell.fill = header_fill; cell.alignment = header_alignment; cell.border = thin_border
            for row_idx, period in enumerate(statistics.get('by_period', []), 2):
                ws_periods.cell(row=row_idx, column=1, value=period['period']).border = thin_border
                ws_periods.cell(row=row_idx, column=2, value=period['avg_days']).border = thin_border
                ws_periods.cell(row=row_idx, column=3, value=period['offers_count']).border = thin_border
                ws_periods.cell(row=row_idx, column=2).alignment = number_alignment
                ws_periods.cell(row=row_idx, column=3).alignment = number_alignment
            for col in range(1, 4):
                ws_periods.column_dimensions[get_column_letter(col)].width = 20

        # Дані кандидатів
        ws_candidates = wb.create_sheet("Дані кандидатів")
        candidate_headers = ["ID", "Ім'я", "Прізвище", "Вакансія", "HR Менеджер", "Дата створення", "Дата офферу", "Днів до офферу"]
        for col, header in enumerate(candidate_headers, 1):
            cell = ws_candidates.cell(row=1, column=col, value=header)
            cell.font = header_font; cell.fill = header_fill; cell.alignment = header_alignment; cell.border = thin_border
        for row_idx, candidate in enumerate(time_data, 2):
            name_parts = candidate['candidate_name'].split() if candidate['candidate_name'] else ['', '']
            ws_candidates.cell(row=row_idx, column=1, value=candidate['candidate_id']).border = thin_border
            ws_candidates.cell(row=row_idx, column=2, value=name_parts[0] if name_parts else '').border = thin_border
            ws_candidates.cell(row=row_idx, column=3, value=' '.join(name_parts[1:]) if len(name_parts) > 1 else '').border = thin_border
            ws_candidates.cell(row=row_idx, column=4, value=candidate['vacancy_title']).border = thin_border
            ws_candidates.cell(row=row_idx, column=5, value=candidate.get('assigned_to_name', '—')).border = thin_border
            ws_candidates.cell(row=row_idx, column=6, value=candidate['new_date'].strftime('%d.%m.%Y') if candidate['new_date'] else '—').border = thin_border
            ws_candidates.cell(row=row_idx, column=7, value=candidate['offer_date'].strftime('%d.%m.%Y') if candidate['offer_date'] else '—').border = thin_border
            ws_candidates.cell(row=row_idx, column=8, value=candidate['days']).border = thin_border
            ws_candidates.cell(row=row_idx, column=8).alignment = number_alignment
        for col in range(1, 9):
            ws_candidates.column_dimensions[get_column_letter(col)].width = 16

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="time_to_hire_{datetime.now().strftime("%Y%m%d")}.xlsx"'
        wb.save(response)
        return response

    # ═══════════════════════════════════════════════════════════
    # TIME-TO-HIRE PDF
    # ═══════════════════════════════════════════════════════════

    @classmethod
    def export_time_to_hire_pdf(cls, time_data, statistics, filters):
        if not REPORTLAB_AVAILABLE:
            return cls._export_time_to_hire_csv_fallback(time_data, statistics, filters)

        try:
            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="time_to_hire_{datetime.now().strftime("%Y%m%d")}.pdf"'

            doc = SimpleDocTemplate(response, pagesize=A4, rightMargin=12*mm, leftMargin=12*mm,
                                   topMargin=15*mm, bottomMargin=15*mm)
            styles = cls._get_styles()
            story = []
            w = doc.width

            # Заголовок
            story.append(Paragraph("Звіт Time-to-Hire", styles['ReportTitle']))
            story.append(Paragraph(f"Сформовано: {timezone.now().strftime('%d.%m.%Y %H:%M:%S')}", styles['ReportSubtitle']))

            # Фільтри
            filter_items = [f"<b>{k}:</b> {v}" for k, v in filters.items() if v]
            if filter_items:
                story.append(Paragraph(" • ".join(filter_items), styles['SmallNote']))
                story.append(Spacer(1, 10))

            # Метрики-картки (3 колонки)
            story.append(Paragraph("Основні метрики", styles['SectionHeader']))
            metric_data = [
                [Paragraph("Середній час", styles['MetricLabel']),
                 Paragraph("Медіана", styles['MetricLabel']),
                 Paragraph("Всього офферів", styles['MetricLabel'])],
                [Paragraph(f"{statistics.get('overall_avg', '—')} <span fontSize=8>днів</span>", styles['MetricValue']),
                 Paragraph(f"{statistics.get('median', '—')} <span fontSize=8>днів</span>", styles['MetricValue']),
                 Paragraph(str(statistics.get('total_offers', 0)), styles['MetricValue'])],
            ]
            metric_table = Table(metric_data, colWidths=[w/3, w/3, w/3])
            metric_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), COLORS['bg']),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('LEFTPADDING', (0, 0), (-1, -1), 6),
                ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                ('LINEBELOW', (0, 0), (-1, 0), 1, COLORS['border']),
            ]))
            story.append(metric_table)
            story.append(Spacer(1, 15))

            # По вакансіях
            if statistics.get('by_vacancy'):
                story.append(Paragraph("Статистика по вакансіях", styles['SectionHeader']))
                vac_data = [[Paragraph("Вакансія", styles['TableHeader']),
                            Paragraph("Сер. час", styles['TableHeader']),
                            Paragraph("Медіана", styles['TableHeader']),
                            Paragraph("Офферів", styles['TableHeader']),
                            Paragraph("Мін", styles['TableHeader']),
                            Paragraph("Макс", styles['TableHeader'])]]
                for v in statistics['by_vacancy']:
                    color = COLORS['success'] if v['avg_days'] <= 14 else COLORS['warning'] if v['avg_days'] <= 30 else COLORS['danger']
                    vac_data.append([
                        Paragraph(v['vacancy_title'][:45], styles['TableCell']),
                        Paragraph(f"<font color={color.hexval()}>{v['avg_days']}</font>", styles['TableCellRight']),
                        Paragraph(str(v['median_days']), styles['TableCellRight']),
                        Paragraph(str(v['offers_count']), styles['TableCellRight']),
                        Paragraph(str(v['min_days']), styles['TableCellRight']),
                        Paragraph(str(v['max_days']), styles['TableCellRight']),
                    ])

                col_w = w / 6
                vac_table = Table(vac_data, colWidths=[col_w*2.5, col_w*0.7, col_w*0.7, col_w*0.7, col_w*0.7, col_w*0.7])
                vac_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), COLORS['header_bg']),
                    ('TEXTCOLOR', (0, 0), (-1, 0), COLORS['white']),
                    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                    ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('GRID', (0, 0), (-1, -1), 0.5, COLORS['border']),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [COLORS['white'], COLORS['bg']]),
                    ('TOPPADDING', (0, 0), (-1, -1), 5),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
                ]))
                story.append(vac_table)
                story.append(Spacer(1, 15))

            # Розподіл
            if statistics.get('distribution'):
                story.append(Paragraph("Розподіл по швидкості", styles['SectionHeader']))
                dist_data = [[Paragraph("Діапазон", styles['TableHeader']),
                             Paragraph("Кількість", styles['TableHeader']),
                             Paragraph("Відсоток", styles['TableHeader'])]]
                dist_colors = [COLORS['success'], COLORS['success'], COLORS['warning'], COLORS['danger'], COLORS['danger']]
                for i, d in enumerate(statistics['distribution']):
                    dist_data.append([
                        Paragraph(d['range'], styles['TableCell']),
                        Paragraph(str(d['count']), styles['TableCellRight']),
                        Paragraph(f"<font color={dist_colors[i].hexval()}>{d['percentage']}%</font>", styles['TableCellRight']),
                    ])
                col_w = w / 3
                dist_table = Table(dist_data, colWidths=[col_w*1.5, col_w*0.75, col_w*0.75])
                dist_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), COLORS['header_bg']),
                    ('TEXTCOLOR', (0, 0), (-1, 0), COLORS['white']),
                    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                    ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('GRID', (0, 0), (-1, -1), 0.5, COLORS['border']),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [COLORS['white'], COLORS['bg']]),
                    ('TOPPADDING', (0, 0), (-1, -1), 5),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
                ]))
                story.append(dist_table)

            doc.build(story)
            return response
        except Exception as e:
            logger.error(f"PDF export error (time_to_hire): {str(e)}\n{traceback.format_exc()}")
            return cls._export_time_to_hire_csv_fallback(time_data, statistics, filters)

    @classmethod
    def _export_time_to_hire_csv_fallback(cls, time_data, statistics, filters):
        import csv
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="time_to_hire_{datetime.now().strftime("%Y%m%d")}.csv"'
        response.write('\ufeff')
        writer = csv.writer(response)
        writer.writerow(['Звіт Time-to-Hire'])
        writer.writerow([f'Дата генерації: {timezone.now().strftime("%d.%m.%Y %H:%M:%S")}'])
        writer.writerow([])
        writer.writerow(['Основні метрики'])
        writer.writerow(['Середній час до офферу (дні)', statistics.get('overall_avg', '—')])
        writer.writerow(['Медіана (дні)', statistics.get('median', '—')])
        writer.writerow(['Всього офферів', statistics.get('total_offers', 0)])
        writer.writerow([])
        writer.writerow(['Деталі по кандидатах'])
        writer.writerow(['ID', "Ім'я", 'Прізвище', 'Вакансія', 'HR Менеджер', 'Дата створення', 'Дата офферу', 'Днів до офферу'])
        for candidate in time_data:
            name_parts = candidate['candidate_name'].split() if candidate['candidate_name'] else ['', '']
            writer.writerow([candidate['candidate_id'], name_parts[0] if name_parts else '', ' '.join(name_parts[1:]) if len(name_parts) > 1 else '', candidate['vacancy_title'], candidate.get('assigned_to_name', '—'), candidate['new_date'].strftime('%d.%m.%Y') if candidate['new_date'] else '—', candidate['offer_date'].strftime('%d.%m.%Y') if candidate['offer_date'] else '—', candidate['days']])
        return response

    # ═══════════════════════════════════════════════════════════
    # HR EFFECTIVENESS EXCEL
    # ═══════════════════════════════════════════════════════════

    @classmethod
    def export_hr_effectiveness_excel(cls, hr_data, summary, filters):
        wb = openpyxl.Workbook()
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1e293b', end_color='1e293b', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        number_alignment = Alignment(horizontal='right', vertical='center')
        center_alignment = Alignment(horizontal='center', vertical='center')

        ws_summary = wb.active
        ws_summary.title = "Загальна статистика"
        ws_summary.merge_cells('A1:D1')
        ws_summary['A1'] = "Звіт Ефективність HR"
        ws_summary['A1'].font = Font(name='Arial', size=18, bold=True, color='1e293b')
        ws_summary['A1'].alignment = center_alignment
        ws_summary['A3'] = "Дата генерації:"
        ws_summary['B3'] = timezone.now().strftime('%d.%m.%Y %H:%M:%S')
        ws_summary['A3'].font = Font(bold=True)

        row = 5
        ws_summary[f'A{row}'] = "Загальні показники"
        ws_summary[f'A{row}'].font = Font(bold=True, size=13, color='1e293b')
        row += 1
        for key, value in [("Кількість HR-менеджерів", summary.get('total_hr', 0)),
                           ("Всього кандидатів", summary.get('total_candidates', 0)),
                           ("Всього офферів", summary.get('total_offers', 0)),
                           ("Загальна конверсія", f"{summary.get('overall_conversion', 0)}%")]:
            ws_summary[f'A{row}'] = key
            ws_summary[f'B{row}'] = value
            ws_summary[f'A{row}'].font = Font(bold=True)
            row += 1

        ws_hr = wb.create_sheet("Деталі по HR")
        hr_headers = ["HR Менеджер", "Email", "Кандидатів", "Офферів", "Співбесід", "Відмов", "Активних", "Конверсія %", "Time-to-Hire"]
        for col, header in enumerate(hr_headers, 1):
            cell = ws_hr.cell(row=1, column=col, value=header)
            cell.font = header_font; cell.fill = header_fill; cell.alignment = header_alignment; cell.border = thin_border
        for row_idx, hr in enumerate(hr_data, 2):
            ws_hr.cell(row=row_idx, column=1, value=hr['hr_name']).border = thin_border
            ws_hr.cell(row=row_idx, column=2, value=hr['hr_email']).border = thin_border
            ws_hr.cell(row=row_idx, column=3, value=hr['total_candidates']).border = thin_border
            ws_hr.cell(row=row_idx, column=4, value=hr['offers_count']).border = thin_border
            ws_hr.cell(row=row_idx, column=5, value=hr['interviews_count']).border = thin_border
            ws_hr.cell(row=row_idx, column=6, value=hr['rejected_count']).border = thin_border
            ws_hr.cell(row=row_idx, column=7, value=hr['active_candidates']).border = thin_border
            ws_hr.cell(row=row_idx, column=8, value=hr['conversion_rate']).border = thin_border
            ws_hr.cell(row=row_idx, column=9, value=hr.get('time_to_hire_avg', '—')).border = thin_border
            for col in range(3, 10):
                if col != 9 or isinstance(ws_hr.cell(row=row_idx, column=col).value, (int, float)):
                    ws_hr.cell(row=row_idx, column=col).alignment = number_alignment
        for col in range(1, 10):
            ws_hr.column_dimensions[get_column_letter(col)].width = 16

        ws_status = wb.create_sheet("Статуси по HR")
        for col, header in enumerate(["HR", "Нові", "Скринінг", "Співбесіда", "Оффер", "Відмова"], 1):
            cell = ws_status.cell(row=1, column=col, value=header)
            cell.font = header_font; cell.fill = header_fill; cell.alignment = header_alignment; cell.border = thin_border
        for row_idx, hr in enumerate(hr_data, 2):
            ws_status.cell(row=row_idx, column=1, value=hr['hr_name']).border = thin_border
            for col_idx, status in enumerate(['new', 'screening', 'interview', 'offer', 'rejected'], 2):
                ws_status.cell(row=row_idx, column=col_idx, value=hr['by_status'].get(status, 0)).border = thin_border
                ws_status.cell(row=row_idx, column=col_idx).alignment = number_alignment

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="hr_effectiveness_{datetime.now().strftime("%Y%m%d")}.xlsx"'
        wb.save(response)
        return response

    # ═══════════════════════════════════════════════════════════
    # HR EFFECTIVENESS PDF — ОХАЙНИЙ З КИРИЛИЦЕЮ
    # ═══════════════════════════════════════════════════════════

    @classmethod
    def export_hr_effectiveness_pdf(cls, hr_data, summary, filters):
        if not REPORTLAB_AVAILABLE:
            return cls._export_hr_effectiveness_csv_fallback(hr_data, summary, filters)

        try:
            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="hr_effectiveness_{datetime.now().strftime("%Y%m%d")}.pdf"'

            doc = SimpleDocTemplate(response, pagesize=A4, rightMargin=12*mm, leftMargin=12*mm,
                                   topMargin=15*mm, bottomMargin=15*mm)
            styles = cls._get_styles()
            story = []
            w = doc.width

            # Заголовок
            story.append(Paragraph("Аналітика ефективності HR-менеджерів", styles['ReportTitle']))
            story.append(Paragraph(f"Звіт сформовано: {timezone.now().strftime('%d.%m.%Y %H:%M')}", styles['ReportSubtitle']))

            # Summary cards
            story.append(Paragraph("Загальні показники", styles['SectionHeader']))
            summary_table_data = [
                [Paragraph("HR-менеджерів", styles['MetricLabel']),
                 Paragraph("Кандидатів", styles['MetricLabel']),
                 Paragraph("Офферів", styles['MetricLabel']),
                 Paragraph("Конверсія", styles['MetricLabel'])],
                [Paragraph(str(summary.get('total_hr', 0)), styles['MetricValue']),
                 Paragraph(str(summary.get('total_candidates', 0)), styles['MetricValue']),
                 Paragraph(str(summary.get('total_offers', 0)), styles['MetricValue']),
                 Paragraph(f"{summary.get('overall_conversion', 0)}%", styles['MetricValue'])],
            ]
            summary_table = Table(summary_table_data, colWidths=[w/4, w/4, w/4, w/4])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), COLORS['bg']),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('LINEBELOW', (0, 0), (-1, 0), 1, COLORS['border']),
            ]))
            story.append(summary_table)
            story.append(Spacer(1, 15))

            # HR таблиця
            story.append(Paragraph("Детальна статистика по HR-менеджерах", styles['SectionHeader']))

            hr_table_data = [[
                Paragraph("HR Менеджер", styles['TableHeader']),
                Paragraph("Канд.", styles['TableHeader']),
                Paragraph("Оффер", styles['TableHeader']),
                Paragraph("Співб.", styles['TableHeader']),
                Paragraph("Відм.", styles['TableHeader']),
                Paragraph("Активн.", styles['TableHeader']),
                Paragraph("Конверс. %", styles['TableHeader']),
            ]]

            for hr in hr_data:
                conv_color = COLORS['success'].hexval() if hr['conversion_rate'] >= 20 else COLORS['warning'].hexval() if hr['conversion_rate'] >= 10 else COLORS['danger'].hexval()
                hr_table_data.append([
                    Paragraph(f"<b>{hr['hr_name']}</b><br/><span fontSize=6 color={COLORS['muted'].hexval()}>@{hr['hr_username']}</span>", styles['TableCell']),
                    Paragraph(str(hr['total_candidates']), styles['TableCellRight']),
                    Paragraph(f"<font color={COLORS['success'].hexval()}>{hr['offers_count']}</font>", styles['TableCellRight']),
                    Paragraph(str(hr['interviews_count']), styles['TableCellRight']),
                    Paragraph(str(hr['rejected_count']), styles['TableCellRight']),
                    Paragraph(str(hr['active_candidates']), styles['TableCellRight']),
                    Paragraph(f"<font color={conv_color}><b>{hr['conversion_rate']}%</b></font>", styles['TableCellRight']),
                ])

            col_w = w / 7
            hr_table = Table(hr_table_data, colWidths=[col_w*2, col_w, col_w, col_w, col_w, col_w, col_w])
            hr_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), COLORS['header_bg']),
                ('TEXTCOLOR', (0, 0), (-1, 0), COLORS['white']),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 0.5, COLORS['border']),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [COLORS['white'], COLORS['bg']]),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('LEFTPADDING', (0, 0), (-1, -1), 4),
                ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ]))
            story.append(hr_table)

            doc.build(story)
            return response
        except Exception as e:
            logger.error(f"PDF export error (hr_effectiveness): {str(e)}\n{traceback.format_exc()}")
            return cls._export_hr_effectiveness_csv_fallback(hr_data, summary, filters)

    @classmethod
    def _export_hr_effectiveness_csv_fallback(cls, hr_data, summary, filters):
        import csv
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="hr_effectiveness_{datetime.now().strftime("%Y%m%d")}.csv"'
        response.write('\ufeff')
        writer = csv.writer(response)
        writer.writerow(['Звіт Ефективність HR'])
        writer.writerow([f'Дата генерації: {timezone.now().strftime("%d.%m.%Y %H:%M:%S")}'])
        writer.writerow([])
        writer.writerow(['Загальні показники'])
        writer.writerow(['Кількість HR-менеджерів', summary.get('total_hr', 0)])
        writer.writerow(['Всього кандидатів', summary.get('total_candidates', 0)])
        writer.writerow(['Всього офферів', summary.get('total_offers', 0)])
        writer.writerow(['Загальна конверсія', f"{summary.get('overall_conversion', 0)}%"])
        writer.writerow([])
        writer.writerow(['Деталі по HR'])
        writer.writerow(['HR Менеджер', 'Email', 'Кандидатів', 'Офферів', 'Співбесід', 'Відмов', 'Активних', 'Конверсія %', 'Time-to-Hire'])
        for hr in hr_data:
            writer.writerow([hr['hr_name'], hr['hr_email'], hr['total_candidates'], hr['offers_count'], hr['interviews_count'], hr['rejected_count'], hr['active_candidates'], hr['conversion_rate'], hr.get('time_to_hire_avg', '—')])
        return response


class FullReportExportService:
    """Сервіс для експорту ПОВНОГО звіту"""

    @classmethod
    def export_full_report_excel(cls, analytics_data):
        wb = openpyxl.Workbook()
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1e293b', end_color='1e293b', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        number_alignment = Alignment(horizontal='right', vertical='center')
        center_alignment = Alignment(horizontal='center', vertical='center')

        # Time-to-Hire
        ws_tth = wb.active
        ws_tth.title = "Time-to-Hire"
        tth_statistics = analytics_data.get('time_to_hire', {})
        ws_tth.merge_cells('A1:F1')
        ws_tth['A1'] = "Звіт Time-to-Hire"
        ws_tth['A1'].font = Font(name='Arial', size=18, bold=True, color='1e293b')
        ws_tth['A1'].alignment = center_alignment
        ws_tth['A3'] = "Дата генерації:"
        ws_tth['B3'] = timezone.now().strftime('%d.%m.%Y %H:%M:%S')
        row = 5
        ws_tth[f'A{row}'] = "Основні метрики"
        ws_tth[f'A{row}'].font = Font(bold=True, size=13, color='1e293b')
        row += 1
        for key, value in [("Середній час до офферу (дні)", tth_statistics.get('overall_avg', '—')),
                           ("Медіана (дні)", tth_statistics.get('median', '—')),
                           ("Всього офферів", tth_statistics.get('total_offers', 0))]:
            ws_tth[f'A{row}'] = key
            ws_tth[f'B{row}'] = value
            row += 1
        if tth_statistics.get('by_vacancy'):
            row += 1
            ws_tth[f'A{row}'] = "Статистика по вакансіях"
            ws_tth[f'A{row}'].font = Font(bold=True, size=13, color='1e293b')
            row += 1
            for col, header in enumerate(["Вакансія", "Середній час", "Медіана", "Офферів", "Мін.", "Макс."], 1):
                cell = ws_tth.cell(row=row, column=col, value=header)
                cell.font = header_font; cell.fill = header_fill; cell.alignment = header_alignment; cell.border = thin_border
            row += 1
            for vacancy in tth_statistics.get('by_vacancy', []):
                for col_idx, key in enumerate(['vacancy_title', 'avg_days', 'median_days', 'offers_count', 'min_days', 'max_days'], 1):
                    ws_tth.cell(row=row, column=col_idx, value=vacancy[key]).border = thin_border
                row += 1

        # HR
        ws_hr = wb.create_sheet("HR-ефективність")
        hr_data = analytics_data.get('hr_effectiveness', {}).get('hr_managers', [])
        hr_summary = analytics_data.get('hr_effectiveness', {}).get('summary', {})
        ws_hr.merge_cells('A1:I1')
        ws_hr['A1'] = "Звіт Ефективність HR"
        ws_hr['A1'].font = Font(name='Arial', size=18, bold=True, color='1e293b')
        ws_hr['A1'].alignment = center_alignment
        ws_hr['A3'] = "Дата генерації:"
        ws_hr['B3'] = timezone.now().strftime('%d.%m.%Y %H:%M:%S')
        row = 5
        ws_hr[f'A{row}'] = "Загальні показники"
        ws_hr[f'A{row}'].font = Font(bold=True, size=13, color='1e293b')
        row += 1
        for key, value in [("Кількість HR-менеджерів", hr_summary.get('total_hr', 0)),
                           ("Всього кандидатів", hr_summary.get('total_candidates', 0)),
                           ("Всього офферів", hr_summary.get('total_offers', 0)),
                           ("Загальна конверсія", f"{hr_summary.get('overall_conversion', 0)}%")]:
            ws_hr[f'A{row}'] = key
            ws_hr[f'B{row}'] = value
            row += 1
        row += 1
        ws_hr[f'A{row}'] = "Деталі по HR-менеджерах"
        ws_hr[f'A{row}'].font = Font(bold=True, size=13, color='1e293b')
        row += 1
        for col, header in enumerate(["HR Менеджер", "Кандидатів", "Офферів", "Співбесід", "Відмов", "Активних", "Конверсія %", "TTH"], 1):
            cell = ws_hr.cell(row=row, column=col, value=header)
            cell.font = header_font; cell.fill = header_fill; cell.alignment = header_alignment; cell.border = thin_border
        row += 1
        for hr in hr_data:
            ws_hr.cell(row=row, column=1, value=hr['hr_name']).border = thin_border
            ws_hr.cell(row=row, column=2, value=hr['total_candidates']).border = thin_border
            ws_hr.cell(row=row, column=3, value=hr['offers_count']).border = thin_border
            ws_hr.cell(row=row, column=4, value=hr['interviews_count']).border = thin_border
            ws_hr.cell(row=row, column=5, value=hr['rejected_count']).border = thin_border
            ws_hr.cell(row=row, column=6, value=hr['active_candidates']).border = thin_border
            ws_hr.cell(row=row, column=7, value=hr['conversion_rate']).border = thin_border
            ws_hr.cell(row=row, column=8, value=hr.get('time_to_hire_avg', '—')).border = thin_border
            row += 1

        # Воронка
        ws_funnel = wb.create_sheet("Воронка кандидатів")
        funnel_data = analytics_data.get('funnel', [])
        total_candidates = analytics_data.get('total_candidates', 1)
        ws_funnel.merge_cells('A1:C1')
        ws_funnel['A1'] = "Воронка кандидатів"
        ws_funnel['A1'].font = Font(name='Arial', size=18, bold=True, color='1e293b')
        ws_funnel['A1'].alignment = center_alignment
        ws_funnel['A3'] = "Дата генерації:"
        ws_funnel['B3'] = timezone.now().strftime('%d.%m.%Y %H:%M:%S')
        ws_funnel['A4'] = "Всього кандидатів:"
        ws_funnel['B4'] = total_candidates
        row = 6
        for col, header in enumerate(["Статус", "Кількість", "Відсоток"], 1):
            cell = ws_funnel.cell(row=row, column=col, value=header)
            cell.font = header_font; cell.fill = header_fill; cell.alignment = header_alignment; cell.border = thin_border
        row += 1
        for status in funnel_data:
            ws_funnel.cell(row=row, column=1, value=status['label']).border = thin_border
            ws_funnel.cell(row=row, column=2, value=status['count']).border = thin_border
            ws_funnel.cell(row=row, column=3, value=f"{status['percentage']}%").border = thin_border
            ws_funnel.cell(row=row, column=2).alignment = number_alignment
            ws_funnel.cell(row=row, column=3).alignment = number_alignment
            row += 1

        # Джерела
        ws_sources = wb.create_sheet("Джерела кандидатів")
        sources_data = analytics_data.get('sources', [])
        ws_sources.merge_cells('A1:C1')
        ws_sources['A1'] = "Кандидати за джерелами"
        ws_sources['A1'].font = Font(name='Arial', size=18, bold=True, color='1e293b')
        ws_sources['A1'].alignment = center_alignment
        ws_sources['A3'] = "Дата генерації:"
        ws_sources['B3'] = timezone.now().strftime('%d.%m.%Y %H:%M:%S')
        row = 5
        for col, header in enumerate(["Джерело", "Кількість", "Відсоток"], 1):
            cell = ws_sources.cell(row=row, column=col, value=header)
            cell.font = header_font; cell.fill = header_fill; cell.alignment = header_alignment; cell.border = thin_border
        row += 1
        for source in sources_data:
            ws_sources.cell(row=row, column=1, value=source['label']).border = thin_border
            ws_sources.cell(row=row, column=2, value=source['count']).border = thin_border
            ws_sources.cell(row=row, column=3, value=f"{source['percentage']}%").border = thin_border
            ws_sources.cell(row=row, column=2).alignment = number_alignment
            ws_sources.cell(row=row, column=3).alignment = number_alignment
            row += 1

        conversion_data = analytics_data.get('source_conversion', [])
        if conversion_data:
            row += 2
            ws_sources[f'A{row}'] = "Конверсія за джерелами"
            ws_sources[f'A{row}'].font = Font(bold=True, size=13, color='1e293b')
            row += 1
            for col, header in enumerate(["Джерело", "Всього", "Оффер %", "Співбесіда+ %"], 1):
                cell = ws_sources.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True); cell.border = thin_border
            row += 1
            for conv in conversion_data:
                ws_sources.cell(row=row, column=1, value=conv['source']).border = thin_border
                ws_sources.cell(row=row, column=2, value=conv['total']).border = thin_border
                ws_sources.cell(row=row, column=3, value=f"{conv['offerRate']}%").border = thin_border
                ws_sources.cell(row=row, column=4, value=f"{conv['interviewRate']}%").border = thin_border
                ws_sources.cell(row=row, column=2).alignment = number_alignment
                row += 1

        # Вакансії
        ws_vacancies = wb.create_sheet("Вакансії")
        vacancies_data = analytics_data.get('vacancies', [])
        ws_vacancies.merge_cells('A1:B1')
        ws_vacancies['A1'] = "Кандидати по вакансіях"
        ws_vacancies['A1'].font = Font(name='Arial', size=18, bold=True, color='1e293b')
        ws_vacancies['A1'].alignment = center_alignment
        ws_vacancies['A3'] = "Дата генерації:"
        ws_vacancies['B3'] = timezone.now().strftime('%d.%m.%Y %H:%M:%S')
        row = 5
        for col, header in enumerate(["Вакансія", "Кількість кандидатів"], 1):
            cell = ws_vacancies.cell(row=row, column=col, value=header)
            cell.font = header_font; cell.fill = header_fill; cell.alignment = header_alignment; cell.border = thin_border
        row += 1
        for vacancy in vacancies_data:
            ws_vacancies.cell(row=row, column=1, value=vacancy['title']).border = thin_border
            ws_vacancies.cell(row=row, column=2, value=vacancy['count']).border = thin_border
            ws_vacancies.cell(row=row, column=2).alignment = number_alignment
            row += 1

        for ws in wb.worksheets:
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[column_letter].width = min(max_length + 2, 35)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="full_analytics_report_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx"'
        wb.save(response)
        return response

    @classmethod
    def export_full_report_pdf(cls, analytics_data):
        if not REPORTLAB_AVAILABLE:
            return cls._export_full_report_csv_fallback(analytics_data)

        try:
            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="full_analytics_report_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf"'

            doc = SimpleDocTemplate(response, pagesize=A4, rightMargin=12*mm, leftMargin=12*mm,
                                   topMargin=15*mm, bottomMargin=15*mm)
            styles = ExportService._get_styles()
            story = []
            w = doc.width

            # Заголовок
            story.append(Paragraph("Повний аналітичний звіт", styles['ReportTitle']))
            story.append(Paragraph(f"Сформовано: {timezone.now().strftime('%d.%m.%Y %H:%M')}", styles['ReportSubtitle']))
            story.append(Spacer(1, 10))

            # 1. Time-to-Hire
            story.append(Paragraph("1. Time-to-Hire", styles['SectionHeader']))
            tth_stats = analytics_data.get('time_to_hire', {})
            tth_data = [[Paragraph("Показник", styles['TableHeader']), Paragraph("Значення", styles['TableHeader'])],
                [Paragraph("Середній час до офферу (дні)", styles['TableCell']), Paragraph(str(tth_stats.get('overall_avg', '—')), styles['TableCellRight'])],
                [Paragraph("Медіана (дні)", styles['TableCell']), Paragraph(str(tth_stats.get('median', '—')), styles['TableCellRight'])],
                [Paragraph("Всього офферів", styles['TableCell']), Paragraph(str(tth_stats.get('total_offers', 0)), styles['TableCellRight'])]]
            tth_table = Table(tth_data, colWidths=[w*0.65, w*0.25])
            tth_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), COLORS['header_bg']),
                ('TEXTCOLOR', (0, 0), (-1, 0), COLORS['white']),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 0.5, COLORS['border']),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [COLORS['white'], COLORS['bg']]),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ]))
            story.append(tth_table)
            story.append(Spacer(1, 15))

            # 2. HR Effectiveness
            story.append(Paragraph("2. Ефективність HR", styles['SectionHeader']))
            hr_summary = analytics_data.get('hr_effectiveness', {}).get('summary', {})
            hr_data = [[Paragraph("Показник", styles['TableHeader']), Paragraph("Значення", styles['TableHeader'])],
                [Paragraph("Кількість HR-менеджерів", styles['TableCell']), Paragraph(str(hr_summary.get('total_hr', 0)), styles['TableCellRight'])],
                [Paragraph("Всього кандидатів", styles['TableCell']), Paragraph(str(hr_summary.get('total_candidates', 0)), styles['TableCellRight'])],
                [Paragraph("Всього офферів", styles['TableCell']), Paragraph(str(hr_summary.get('total_offers', 0)), styles['TableCellRight'])],
                [Paragraph("Загальна конверсія", styles['TableCell']), Paragraph(f"{hr_summary.get('overall_conversion', 0)}%", styles['TableCellRight'])]]
            hr_table = Table(hr_data, colWidths=[w*0.65, w*0.25])
            hr_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), COLORS['header_bg']),
                ('TEXTCOLOR', (0, 0), (-1, 0), COLORS['white']),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 0.5, COLORS['border']),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [COLORS['white'], COLORS['bg']]),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ]))
            story.append(hr_table)
            story.append(Spacer(1, 15))

            # 3. Воронка
            story.append(Paragraph("3. Воронка кандидатів", styles['SectionHeader']))
            funnel_data = analytics_data.get('funnel', [])
            if funnel_data:
                funnel_table_data = [[Paragraph("Статус", styles['TableHeader']),
                                    Paragraph("Кількість", styles['TableHeader']),
                                    Paragraph("Відсоток", styles['TableHeader'])]]
                for f in funnel_data:
                    funnel_table_data.append([Paragraph(f['label'], styles['TableCell']),
                                             Paragraph(str(f['count']), styles['TableCellRight']),
                                             Paragraph(f"{f['percentage']}%", styles['TableCellRight'])])
                col_w = w / 3
                funnel_table = Table(funnel_table_data, colWidths=[col_w*1.5, col_w*0.75, col_w*0.75])
                funnel_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), COLORS['header_bg']),
                    ('TEXTCOLOR', (0, 0), (-1, 0), COLORS['white']),
                    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                    ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('GRID', (0, 0), (-1, -1), 0.5, COLORS['border']),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [COLORS['white'], COLORS['bg']]),
                    ('TOPPADDING', (0, 0), (-1, -1), 5),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
                ]))
                story.append(funnel_table)

            doc.build(story)
            return response
        except Exception as e:
            logger.error(f"PDF export error (full_report): {str(e)}\n{traceback.format_exc()}")
            return cls._export_full_report_csv_fallback(analytics_data)

    @classmethod
    def _export_full_report_csv_fallback(cls, analytics_data):
        import csv
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="full_analytics_report_{datetime.now().strftime("%Y%m%d_%H%M")}.csv"'
        response.write('\ufeff')
        writer = csv.writer(response)
        writer.writerow(['ПОВНИЙ АНАЛІТИЧНИЙ ЗВІТ'])
        writer.writerow([f'Дата генерації: {timezone.now().strftime("%d.%m.%Y %H:%M:%S")}'])
        writer.writerow([])
        writer.writerow(['=== TIME-TO-HIRE ==='])
        tth_stats = analytics_data.get('time_to_hire', {})
        writer.writerow(['Середній час до офферу (дні)', tth_stats.get('overall_avg', '—')])
        writer.writerow(['Медіана (дні)', tth_stats.get('median', '—')])
        writer.writerow(['Всього офферів', tth_stats.get('total_offers', 0)])
        writer.writerow([])
        writer.writerow(['=== ЕФЕКТИВНІСТЬ HR ==='])
        hr_summary = analytics_data.get('hr_effectiveness', {}).get('summary', {})
        writer.writerow(['Кількість HR-менеджерів', hr_summary.get('total_hr', 0)])
        writer.writerow(['Всього кандидатів', hr_summary.get('total_candidates', 0)])
        writer.writerow(['Всього офферів', hr_summary.get('total_offers', 0)])
        writer.writerow(['Загальна конверсія', f"{hr_summary.get('overall_conversion', 0)}%"])
        writer.writerow([])
        writer.writerow(['=== ВОРОНКА КАНДИДАТІВ ==='])
        writer.writerow(['Статус', 'Кількість', 'Відсоток'])
        for f in analytics_data.get('funnel', []):
            writer.writerow([f['label'], f['count'], f"{f['percentage']}%"])
        return response